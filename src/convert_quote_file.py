#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from docx import Document
from openpyxl import Workbook

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional in local fallback environments
    PdfReader = None  # type: ignore[assignment]


STANDARD_HEADERS = ["追溯号", "名称", "规格型号", "单位", "数量", "品牌/厂家", "税率", "供货周期", "付款条件", "报价单价", "报价合价", "报价备注"]
ALIASES = {
    "追溯号": ["追溯号", "内部追溯号", "原始行号", "编号"],
    "名称": ["名称", "材料名称", "标准材料名称", "项目名称", "品名"],
    "规格型号": ["规格型号", "规格", "型号", "规格及型号"],
    "单位": ["单位"],
    "数量": ["数量", "工程量"],
    "品牌/厂家": ["品牌/厂家", "品牌", "厂家", "制造商"],
    "税率": ["税率"],
    "供货周期": ["供货周期", "工期", "交货期"],
    "付款条件": ["付款条件", "付款方式"],
    "报价单价": ["报价单价", "单价", "综合单价", "含税单价"],
    "报价合价": ["报价合价", "合价", "总价", "金额", "含税合价"],
    "报价备注": ["报价备注", "备注", "说明"],
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).replace("\n", " ").strip()


def normalize(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def looks_like_header(row: list[str]) -> bool:
    text = "|".join(normalize(cell) for cell in row)
    hits = 0
    for aliases in ALIASES.values():
        if any(normalize(alias) in text for alias in aliases):
            hits += 1
    return hits >= 2


def column_map(row: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, cell in enumerate(row):
        norm = normalize(cell)
        for header, aliases in ALIASES.items():
            if any(normalize(alias) == norm or normalize(alias) in norm for alias in aliases):
                mapping[idx] = header
                break
    return mapping


def rows_from_table(table_rows: list[list[str]]) -> list[dict[str, str]]:
    header_idx = next((idx for idx, row in enumerate(table_rows[:8]) if looks_like_header(row)), None)
    if header_idx is None:
        return []
    mapping = column_map(table_rows[header_idx])
    rows: list[dict[str, str]] = []
    for source in table_rows[header_idx + 1:]:
        item = {header: "" for header in STANDARD_HEADERS}
        for idx, value in enumerate(source):
            header = mapping.get(idx)
            if header:
                item[header] = clean(value)
        if item["名称"] or item["追溯号"] or item["报价单价"] or item["报价合价"]:
            rows.append(item)
    return rows


def extract_docx(path: Path) -> tuple[list[dict[str, str]], list[list[str]]]:
    doc = Document(path)
    raw_tables: list[list[str]] = []
    rows: list[dict[str, str]] = []
    for table in doc.tables:
        table_rows = [[clean(cell.text) for cell in row.cells] for row in table.rows]
        raw_tables.extend(table_rows)
        rows.extend(rows_from_table(table_rows))
    return rows, raw_tables


def split_pdf_line(line: str) -> list[str]:
    parts = [part.strip() for part in re.split(r"\s{2,}|\t+", line) if part.strip()]
    return parts if len(parts) > 1 else [line.strip()]


def extract_pdf(path: Path) -> tuple[list[dict[str, str]], list[list[str]]]:
    if PdfReader is None:
        return [], [["未安装 pypdf，无法读取 PDF。"]]
    reader = PdfReader(str(path))
    raw_rows: list[list[str]] = []
    for page_no, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        raw_rows.append([f"第{page_no}页"])
        for line in text.splitlines():
            line = line.strip()
            if line:
                raw_rows.append(split_pdf_line(line))
    rows: list[dict[str, str]] = []
    window: list[list[str]] = []
    for raw in raw_rows:
        if looks_like_header(raw):
            window = [raw]
            continue
        if window:
            window.append(raw)
    if window:
        rows = rows_from_table(window)
    if not rows:
        for raw in raw_rows:
            text = " ".join(raw)
            if text and not text.startswith("第"):
                rows.append({**{header: "" for header in STANDARD_HEADERS}, "名称": text, "报价备注": "PDF自动提取文本，需人工确认字段对应关系"})
    return rows, raw_rows


def write_workbook(output: Path, source: Path, rows: list[dict[str, str]], raw_rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "正式报价清单"
    headers = ["序号", *STANDARD_HEADERS]
    ws.append(headers)
    for idx, row in enumerate(rows, 1):
        ws.append([idx, *[row.get(header, "") for header in STANDARD_HEADERS]])

    raw_ws = wb.create_sheet("原始提取内容")
    raw_ws.append(["来源文件", str(source)])
    raw_ws.append(["说明", "此表为 Word/PDF 自动提取结果；若正式报价清单字段不完整，请人工核对后再分析。"])
    raw_ws.append([])
    for row in raw_rows:
        raw_ws.append(row)

    note_ws = wb.create_sheet("转换说明")
    note_ws.append(["事项", "说明"])
    note_ws.append(["转换状态", "已生成正式报价清单" if rows else "未识别到有效报价行"])
    note_ws.append(["识别报价行数", len(rows)])
    note_ws.append(["人工确认要求", "重点核对追溯号、名称、规格型号、数量、单价、合价是否对应原报价文件。"])
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="将供应商 Word/PDF 报价单预转换为可分析 Excel")
    parser.add_argument("--input", required=True, help="供应商报价源文件，支持 .docx/.pdf")
    parser.add_argument("--output", required=True, help="输出 xlsx 文件")
    args = parser.parse_args()
    source = Path(args.input).resolve()
    output = Path(args.output).resolve()
    if source.suffix.lower() == ".docx":
        rows, raw_rows = extract_docx(source)
    elif source.suffix.lower() == ".pdf":
        rows, raw_rows = extract_pdf(source)
    else:
        raise RuntimeError("仅支持 .docx/.pdf 转换")
    write_workbook(output, source, rows, raw_rows)
    print(f"源文件：{source}")
    print(f"识别报价行：{len(rows)}")
    print(f"输出文件：{output}")


if __name__ == "__main__":
    main()
