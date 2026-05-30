#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = TOOL_DIR.parent

STANDARD_HEADERS = [
    "序号", "来源行号", "原始编号", "原始材料名称", "标准材料名称", "询价大类", "询价采购包", "供应商类型",
    "规格型号", "厚度参数", "参数选项", "单位", "合并数量", "原始参考单价", "原始参考合价",
    "是否需人工确认", "是否可直接发供应商询价", "询价前需补充信息", "备注",
    "供应商报价单价", "供应商报价合价", "品牌/厂家", "税率", "供货周期", "付款条件", "报价备注",
]

OUTPUT_HEADERS = ["追溯号"] + STANDARD_HEADERS


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def find_header_row(ws) -> tuple[int, list[str]]:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [clean(cell.value) for cell in ws[row_idx]]
        joined = "|".join(values)
        if any(key in joined for key in ["材料名称", "原始材料名称", "标准材料名称", "名称"]) and any(key in joined for key in ["单位", "数量", "合并数量"]):
            return row_idx, values
    return 1, [clean(cell.value) for cell in ws[1]]


def pick(row: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if name in row and clean(row.get(name)):
            return row.get(name)
    return ""


def normalize_headers(headers: list[str]) -> list[str]:
    return [clean(header) or f"空列{idx}" for idx, header in enumerate(headers, 1)]


def spec_from_text(name: str, existing: str) -> str:
    if existing:
        return existing
    patterns = [
        r"BTLY[-\w×+\.]+mm2?",
        r"WDZ[N]?-?[A-Z]+[-\w×+\.]*\d+(?:\.\d+)?mm2?",
        r"YJY[-\w×+\.]*\d+(?:\.\d+)?mm2?",
        r"BV[R]?-?\d+(?:\.\d+)?mm2?",
        r"DN\s*\d+",
        r"Φ\s*\d+(?:[-~]\d+)?",
        r"M\d+(?:[×xX]\d+(?:-\d+)?)?",
        r"\d+(?:\.\d+)?\s*(?:厚|mm|cm)",
        r"\d+\s*[×xX*]\s*\d+(?:\s*[×xX*]\s*\d+)?",
        r"C\d+",
        r"[甲乙丙]级",
    ]
    for pattern in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            return match.group(0).replace("*", "×").replace("x", "×").replace("X", "×")
    return ""


def strip_spec(name: str, spec: str) -> str:
    if spec and spec in name and len(name) > len(spec):
        return clean(name.replace(spec, ""))
    return name


def classify(name: str, spec: str, unit: str, original_category: str = "") -> tuple[str, str, str, str, str, str]:
    text = f"{name} {spec} {original_category}"
    manual = "否"
    direct = "是"
    need = ""

    def result(category: str, package: str, supplier: str, direct_default: str = "是", need_default: str = ""):
        nonlocal manual, direct, need
        direct = direct_default
        if need_default:
            need = need_default
        if not spec and direct_default == "是":
            manual = "是"
            direct = "否"
            need = "需补充：规格型号、人工确认"
        return category, package, supplier, manual, direct, need

    if any(k in text for k in ["WDZ", "WDZN", "YJY", "BYJ", "BTLY", "BVR", "电缆", "电线"]):
        return result("电气电线电缆灯具类", "电线电缆包", "电线电缆供应商")
    if any(k in text for k in ["配电箱", "控制箱", "柜", "充电枪配电"]):
        return result("电气电线电缆灯具类", "配电箱柜包", "成套电气设备供应商", "否", "需补充：系统图、箱内元器件配置、规格型号、人工确认")
    if any(k in text for k in ["桥架", "线管", "电缆托架", "穿线管"]):
        return result("电气电线电缆灯具类", "桥架线管辅材包", "电气管线桥架供应商")
    if any(k in text for k in ["灯", "照明", "开关", "插座"]):
        return result("电气电线电缆灯具类", "灯具照明包", "灯具照明供应商", "否", "需补充：品牌、功率、色温、防护等级或安装方式")
    if any(k in text for k in ["消防报警", "模块", "火灾", "总线", "多线", "应急照明控制器", "UPS"]):
        return result("弱电消防报警类", "消防报警及弱电设备包", "消防报警弱电设备供应商", "否", "需补充：品牌系统、接口、编码、调试责任")
    if any(k in text for k in ["防火阀", "风阀", "风口", "风管", "百叶", "排烟"]):
        return result("暖通通风类", "暖通风管风阀风口包", "暖通风管及阀部件供应商", "否", "需补充：关键参数；材质、板厚、连接方式、执行机构、防火/排烟要求需确认")
    if any(k in text for k in ["风机", "排气扇", "除臭"]):
        return result("暖通通风类", "暖通通风设备包", "暖通设备供应商", "否", "需补充：设备参数、控制方式、安装调试范围")
    if any(k in text for k in ["消火栓", "喷淋", "灭火器", "消防水泵接合器", "消防箱", "水流指示器", "末端试水"]):
        return result("消防给水及消防设备类", "消防给水管材阀门附件包", "消防水系统管材阀门供应商", "否", "需补充：消防系统参数、连接方式、验收要求")
    if any(k in text for k in ["水泵", "潜污泵", "稳压设备", "加压泵"]):
        return result("消防给水及消防设备类", "消防给水设备包", "消防水系统设备供应商", "否", "需补充：设备流量、扬程、功率、控制柜、安装调试范围")
    if any(k in text for k in ["DN", "管", "阀", "弯头", "三通", "法兰", "接头", "水表", "洁具", "地漏"]):
        return result("给排水管材管件阀门洁具类", "零星待确认包", "待采购员确认供应商类型", "否", "需确认给排水/消防/暖通归属及参数")
    if any(k in text for k in ["钢筋", "钢板", "型钢", "角钢", "槽钢", "方钢", "钢管", "镀锌", "不锈钢", "铝合金板"]):
        return result("钢材及金属型材类", "钢材型材采购包", "钢材型材供应商")
    if any(k in text for k in ["螺栓", "螺钉", "螺母", "垫圈", "膨胀螺丝", "铆钉"]):
        return result("五金紧固件及辅材类", "五金紧固件辅材包", "五金辅材供应商")
    if any(k in text for k in ["防水", "卷材", "沥青", "聚氨酯", "保温", "聚苯", "密封"]):
        return result("防水保温密封类", "防水保温密封材料包", "防水保温材料供应商")
    if any(k in text for k in ["门", "窗", "玻璃", "栏杆"]):
        return result("门窗玻璃栏杆类", "门窗玻璃栏杆包", "门窗栏杆专业供应商", "否" if not spec else "是", "需补充门窗表、节点、五金、安装边界" if not spec else "")
    if any(k in text for k in ["混凝土"]):
        package = "商品混凝土包" if "商品" in text else "混凝土砂浆材料包"
        return result("混凝土及砂浆类", package, "商品混凝土供应商" if package == "商品混凝土包" else "混凝土砂浆材料供应商")
    if any(k in text for k in ["砂浆", "外加剂", "膨胀剂"]):
        return result("混凝土及砂浆类", "砂浆外加剂包", "砂浆及外加剂供应商")
    if "水泥" in text:
        return result("水泥砂石砖砌体类", "水泥包", "水泥供应商")
    if any(k in text for k in ["砂", "石", "砖", "砌块", "瓦"]):
        package = "砖砌块包" if any(k in text for k in ["砖", "砌块"]) else "砂石骨料包"
        return result("水泥砂石砖砌体类", package, "砖砌块供应商" if package == "砖砌块包" else "砂石骨料供应商")
    if any(k in text for k in ["木", "模板", "脚手架", "扣件", "可调托座"]):
        return result("木材模板脚手架周转材料类", "模板木材脚手架包", "模板木材周转材料供应商")
    if any(k in text for k in ["油漆", "涂料", "胶", "丙酮", "乙醇", "稀释剂", "固化剂", "乙炔", "氧气"]):
        return result("油漆胶粘剂化工辅材类", "油漆胶粘剂化工辅材包", "油漆化工辅材供应商")
    if any(k in text for k in ["瓷砖", "面砖", "地板", "龙骨", "石膏板", "装饰"]):
        return result("装饰装修材料类", "装饰装修材料包", "装饰材料供应商")
    if any(k in text for k in ["安全", "标志牌", "防坠网", "养护膜", "土工布", "减速带"]):
        return result("安全文明及临时措施材料类", "安全文明临时措施材料包", "安全文明材料供应商")

    return "零星材料/需人工确认类", "零星待确认包", "待采购员确认供应商类型", "是", "否", "需补充：人工确认"


def source_rows(path: Path) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except BadZipFile as exc:
        raise RuntimeError("当前文件不是有效的 .xlsx 工作簿。若原文件是 .xls，请先用 Excel/WPS 另存为 .xlsx 后重新上传。") from exc
    if "正式询价清单" in wb.sheetnames:
        ws = wb["正式询价清单"]
    elif "正式询价清单底表" in wb.sheetnames:
        ws = wb["正式询价清单底表"]
    else:
        ws = wb[wb.sheetnames[0]]
    header_row, headers = find_header_row(ws)
    headers = normalize_headers(headers)
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(headers, values))
        if any(clean(value) for value in values):
            rows.append(row)
    return rows


def transform(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    transformed: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, 1):
        if all(header in row for header in STANDARD_HEADERS[:8]):
            out = {header: row.get(header, "") for header in STANDARD_HEADERS}
            out["追溯号"] = f"{clean(out.get('来源行号'))}｜{clean(out.get('原始编号'))}" if clean(out.get("原始编号")) else clean(out.get("来源行号"))
            transformed.append(out)
            continue

        source_row = clean(pick(row, ["来源行号", "原始行号", "行号", "序号"])) or str(idx)
        code = clean(pick(row, ["原始编号", "编号", "材料编码", "编码"]))
        original_name = clean(pick(row, ["原始材料名称", "材料名称", "名称", "材料设备名称", "项目名称"]))
        raw_spec = clean(pick(row, ["规格型号", "规格", "型号", "项目特征"]))
        unit = clean(pick(row, ["单位", "计量单位"]))
        qty = pick(row, ["合并数量", "数量", "工程量", "消耗量"])
        unit_price = pick(row, ["原始参考单价", "参考单价", "单价", "市场价"])
        total = pick(row, ["原始参考合价", "参考合价", "合价", "金额"])
        original_category = clean(pick(row, ["原始类别", "类别", "材料大类"]))
        spec = spec_from_text(original_name, raw_spec)
        std_name = strip_spec(original_name, spec)
        category, package, supplier, manual, direct, need = classify(original_name, spec, unit, original_category)
        out = {
            "追溯号": f"{source_row}｜{code}" if code else source_row,
            "序号": idx,
            "来源行号": source_row,
            "原始编号": code,
            "原始材料名称": original_name,
            "标准材料名称": std_name,
            "询价大类": category,
            "询价采购包": package,
            "供应商类型": supplier,
            "规格型号": spec,
            "厚度参数": spec if "厚" in spec else "",
            "参数选项": "",
            "单位": unit,
            "合并数量": qty,
            "原始参考单价": unit_price,
            "原始参考合价": total,
            "是否需人工确认": manual,
            "是否可直接发供应商询价": direct,
            "询价前需补充信息": need,
            "备注": f"原类别：{original_category}" if original_category else "",
            "供应商报价单价": "",
            "供应商报价合价": "",
            "品牌/厂家": "",
            "税率": "",
            "供货周期": "",
            "付款条件": "",
            "报价备注": "",
        }
        transformed.append(out)
    return transformed


def style_sheet(ws, header_row: int = 3) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    blue = PatternFill("solid", fgColor="5B9BD5")
    orange = PatternFill("solid", fgColor="FCE4D6")
    green = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    manual_col = direct_col = None
    for idx, cell in enumerate(ws[header_row], 1):
        if cell.value == "是否需人工确认":
            manual_col = idx
        if cell.value == "是否可直接发供应商询价":
            direct_col = idx
    for row in ws.iter_rows(min_row=header_row + 1):
        if manual_col and row[manual_col - 1].value == "是":
            for cell in row:
                cell.fill = orange
        elif direct_col and row[direct_col - 1].value == "是":
            for cell in row:
                cell.fill = green
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.sheet_view.showGridLines = False
    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_sheet(wb: Workbook, name: str, title: str, headers: list[str], data: list[list[Any]], widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title)
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    for row_idx, row in enumerate(data, 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
    style_sheet(ws)
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)


def write_output(path: Path, rows: list[dict[str, Any]], source_file: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    package_counter = Counter(row["询价采购包"] for row in rows)
    manual_rows = [row for row in rows if row["是否需人工确认"] == "是"]
    direct_rows = [row for row in rows if row["是否可直接发供应商询价"] == "是"]
    summary = [
        ["待处理清单", str(source_file)],
        ["清单行数", len(rows)],
        ["采购包数量", len(package_counter)],
        ["可直接询价项", len(direct_rows)],
        ["需人工确认项", len(manual_rows)],
        ["说明", "本文件由未分类待处理清单自动识别生成，分类规则需在实际使用中继续校准。"],
    ]
    for package, count in package_counter.most_common():
        summary.append([f"采购包：{package}", count])
    write_sheet(wb, "处理说明", "处理说明", ["项目", "内容"], summary, {"项目": 38, "内容": 90})
    widths = {
        "追溯号": 24, "序号": 8, "来源行号": 12, "原始编号": 20, "原始材料名称": 38, "标准材料名称": 38,
        "询价大类": 30, "询价采购包": 32, "供应商类型": 34, "规格型号": 32, "参数选项": 50,
        "单位": 10, "合并数量": 14, "是否需人工确认": 16, "是否可直接发供应商询价": 20,
        "询价前需补充信息": 70, "备注": 40,
    }
    matrix = [[row.get(header, "") for header in OUTPUT_HEADERS] for row in rows]
    write_sheet(wb, "采购包拆分总表", "采购包拆分总表", OUTPUT_HEADERS, matrix, widths)
    write_sheet(wb, "可直接询价项", "可直接询价项", OUTPUT_HEADERS, [[row.get(header, "") for header in OUTPUT_HEADERS] for row in direct_rows], widths)
    write_sheet(wb, "需人工确认项", "需人工确认项", OUTPUT_HEADERS, [[row.get(header, "") for header in OUTPUT_HEADERS] for row in manual_rows], widths)
    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="处理未分类采购询价清单并生成采购包拆分总表")
    parser.add_argument("--base-dir", default=str(PROJECT_DIR), help="项目根目录")
    parser.add_argument("--input", required=True, help="待处理清单 xlsx")
    parser.add_argument("--output-dir", default="采购询价本地工作流/output", help="输出目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir).resolve()
    input_path = resolve_path(base_dir, args.input)
    output_root = resolve_path(base_dir, args.output_dir)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / f"{run_id}_由待处理清单生成拆分"
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = transform(source_rows(input_path))
    output_path = output_dir / "采购包拆分总表-由待处理清单生成.xlsx"
    write_output(output_path, rows, input_path)
    log = [
        f"运行时间：{run_id}",
        f"待处理清单：{input_path}",
        f"输出目录：{output_dir}",
        f"清单行数：{len(rows)}",
        f"采购包数量：{len(set(row['询价采购包'] for row in rows))}",
        f"可直接询价项：{sum(1 for row in rows if row['是否可直接发供应商询价'] == '是')}",
        f"需人工确认项：{sum(1 for row in rows if row['是否需人工确认'] == '是')}",
        f"输出文件：{output_path}",
    ]
    (output_dir / "处理日志.txt").write_text("\n".join(log) + "\n", encoding="utf-8")
    print("\n".join(log))


if __name__ == "__main__":
    main()
