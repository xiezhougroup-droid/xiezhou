#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = TOOL_DIR.parent

QUOTE_SHEET = "正式报价清单"
QUOTE_HEADERS = [
    "序号",
    "来源行号",
    "追溯号",
    "名称",
    "规格型号",
    "单位",
    "数量",
    "品牌/厂家",
    "税率",
    "供货周期",
    "付款条件",
    "报价单价",
    "报价合价",
    "报价备注",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def supplier_name_from_path(path: Path) -> str:
    stem = path.stem
    for suffix in ["-报价清单", "报价清单", "-报价", "报价"]:
        stem = stem.replace(suffix, "")
    return stem.strip(" -_") or path.stem


def read_quote_file(path: Path, package_hint: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if QUOTE_SHEET not in wb.sheetnames:
        return []
    ws = wb[QUOTE_SHEET]
    header_row = None
    headers = None
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [cell.value for cell in ws[row_idx]]
        if "追溯号" in values and "报价单价" in values:
            header_row = row_idx
            headers = values
            break
    if header_row is None or headers is None:
        return []
    supplier = supplier_name_from_path(path)
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(headers, values))
        trace = clean(row.get("追溯号"))
        name = clean(row.get("名称"))
        if not trace and not name:
            continue
        qty = number(row.get("数量"))
        unit_price = number(row.get("报价单价"))
        total_price = number(row.get("报价合价"))
        if total_price is None and qty is not None and unit_price is not None:
            total_price = qty * unit_price
        rows.append({
            "供应商": supplier,
            "文件名": path.name,
            "采购包": package_hint or path.parent.name,
            "序号": row.get("序号"),
            "来源行号": clean(row.get("来源行号")),
            "追溯号": trace,
            "名称": name,
            "规格型号": clean(row.get("规格型号")),
            "单位": clean(row.get("单位")),
            "数量": qty,
            "品牌/厂家": clean(row.get("品牌/厂家")),
            "税率": clean(row.get("税率")),
            "供货周期": clean(row.get("供货周期")),
            "付款条件": clean(row.get("付款条件")),
            "报价单价": unit_price,
            "报价合价": total_price,
            "报价备注": clean(row.get("报价备注")),
        })
    return rows


def collect_quotes(input_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in sorted(input_dir.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        rows.extend(read_quote_file(path))
    return rows


def strategy_for_supplier(summary: dict[str, Any], best_total: float | None) -> str:
    if summary["已报价项"] == 0:
        return "未形成有效报价，先要求按原清单补齐报价。"
    points = []
    if summary["漏报价项"] > 0:
        points.append(f"先要求补齐 {summary['漏报价项']} 个漏报价项，并确认是否有清单外限制条件")
    if best_total and summary["报价总价"] and summary["报价总价"] > best_total * 1.05:
        points.append("总价高于最低价超过5%，要求针对高价项二次报价")
    if summary["异常项"] > 0:
        points.append("对单价明显偏高/偏低或缺少品牌、税率、周期的异常项逐项澄清")
    if not points:
        points.append("报价完整度较好，可要求锁定价格、供货周期、税率和付款条件")
    return "；".join(points) + "。"


def style_sheet(ws, header_row: int = 3) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    blue = PatternFill("solid", fgColor="5B9BD5")
    orange = PatternFill("solid", fgColor="FCE4D6")
    green = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if ws.max_row >= header_row:
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [clean(cell.value) for cell in row]
        if any("异常" in value or "漏" in value or "需" in value for value in values):
            for cell in row:
                cell.fill = orange
        elif any("推荐" in value or "完整" in value for value in values):
            for cell in row:
                cell.fill = green
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.sheet_view.showGridLines = False
    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_sheet(wb: Workbook, name: str, title: str, headers: list[str], rows: list[list[Any]], widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws.cell(1, 1, title)
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    style_sheet(ws)
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)


def analyze(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    by_package_supplier: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_package_supplier[(row["采购包"], row["供应商"])].append(row)

    expected_by_package: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row["追溯号"]:
            expected_by_package[row["采购包"]].add(row["追溯号"])

    summaries: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    for (package, supplier), items in by_package_supplier.items():
        quoted_items = [item for item in items if item["报价单价"] is not None or item["报价合价"] is not None]
        total = sum(item["报价合价"] or 0 for item in quoted_items)
        expected = expected_by_package[package]
        quoted_trace = {item["追溯号"] for item in quoted_items if item["追溯号"]}
        missing = len(expected - quoted_trace)
        item_prices = [item["报价单价"] for item in quoted_items if item["报价单价"] is not None]
        avg_price = sum(item_prices) / len(item_prices) if item_prices else None
        issue_count = 0
        for item in items:
            issue = []
            if item["报价单价"] is None and item["报价合价"] is None:
                issue.append("未报价")
            if not item["品牌/厂家"]:
                issue.append("缺品牌/厂家")
            if not item["税率"]:
                issue.append("缺税率")
            if not item["供货周期"]:
                issue.append("缺供货周期")
            if avg_price and item["报价单价"] and (item["报价单价"] > avg_price * 1.8 or item["报价单价"] < avg_price * 0.4):
                issue.append("单价偏离较大")
            if issue:
                issue_count += 1
                anomalies.append({**item, "异常说明": "；".join(issue)})
        summaries.append({
            "采购包": package,
            "供应商": supplier,
            "清单项数": len(expected),
            "已报价项": len(quoted_trace),
            "漏报价项": missing,
            "报价总价": total,
            "异常项": issue_count,
            "完整度": len(quoted_trace) / len(expected) if expected else 0,
        })
    best_by_package: dict[str, float] = {}
    for summary in summaries:
        total = summary["报价总价"]
        if total:
            best_by_package[summary["采购包"]] = min(best_by_package.get(summary["采购包"], total), total)
    for summary in summaries:
        best = best_by_package.get(summary["采购包"])
        summary["总价排名参考"] = "最低价" if best and summary["报价总价"] == best else ""
        summary["谈判策略"] = strategy_for_supplier(summary, best)

    recommendations: list[dict[str, Any]] = []
    for package in sorted({summary["采购包"] for summary in summaries}):
        candidates = [summary for summary in summaries if summary["采购包"] == package and summary["已报价项"] > 0]
        candidates.sort(key=lambda x: (x["漏报价项"], x["报价总价"] or 10**18, -x["完整度"], x["异常项"]))
        if candidates:
            best = candidates[0]
            recommendations.append({
                "采购包": package,
                "阶段性推荐供应商": best["供应商"],
                "推荐理由": f"报价完整度 {best['完整度']:.0%}，漏报价 {best['漏报价项']} 项，当前总价 {best['报价总价']:.2f}。",
                "风险提示": "需完成异常项澄清、商务条件确认和采购员谈判反馈后再形成最终合作建议。",
            })
    return summaries, anomalies, recommendations


def write_analysis(output_path: Path, input_dir: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    if not rows:
        write_sheet(
            wb,
            "使用说明",
            "供应商报价分析使用说明",
            ["事项", "说明"],
            [
                ["当前状态", "未在输入目录中发现可分析的供应商报价 Excel。"],
                ["报价文件放置位置", str(input_dir)],
                ["报价文件要求", "把供应商返回的报价清单 xlsx 放入 input/supplier_quotes/采购包名称/ 文件夹，文件名建议包含供应商名称。"],
                ["识别依据", "脚本读取“正式报价清单”sheet，并通过“追溯号”识别报价项。"],
                ["下一步", "收齐供应商报价后重新运行 run_analyze_quotes.command。"],
            ],
            {"事项": 28, "说明": 100},
        )
        wb.save(output_path)
        return

    summaries, anomalies, recommendations = analyze(rows)
    write_sheet(
        wb,
        "处理说明",
        "供应商报价分析处理说明",
        ["项目", "内容"],
        [
            ["输入目录", str(input_dir)],
            ["读取报价行", len(rows)],
            ["供应商-采购包组合", len(summaries)],
            ["异常项", len(anomalies)],
            ["说明", "本结果为自动分析初稿，最终推荐需采购员确认谈判反馈、技术偏差和商务条件。"],
        ],
        {"项目": 30, "内容": 100},
    )
    summary_headers = ["采购包", "供应商", "清单项数", "已报价项", "漏报价项", "报价总价", "异常项", "完整度", "总价排名参考", "谈判策略"]
    write_sheet(
        wb,
        "报价汇总及谈判策略",
        "报价汇总及谈判策略",
        summary_headers,
        [[s[h] for h in summary_headers] for s in summaries],
        {"采购包": 28, "供应商": 26, "谈判策略": 80},
    )
    detail_headers = ["采购包", "供应商", "追溯号", "名称", "规格型号", "单位", "数量", "品牌/厂家", "税率", "供货周期", "付款条件", "报价单价", "报价合价", "报价备注"]
    write_sheet(
        wb,
        "报价明细合并表",
        "报价明细合并表",
        detail_headers,
        [[row.get(h, "") for h in detail_headers] for row in rows],
        {"采购包": 28, "供应商": 26, "追溯号": 24, "名称": 38, "规格型号": 34, "报价备注": 30},
    )
    anomaly_headers = detail_headers + ["异常说明"]
    write_sheet(
        wb,
        "异常及澄清项",
        "异常及澄清项",
        anomaly_headers,
        [[row.get(h, "") for h in anomaly_headers] for row in anomalies],
        {"采购包": 28, "供应商": 26, "追溯号": 24, "名称": 38, "规格型号": 34, "异常说明": 50},
    )
    rec_headers = ["采购包", "阶段性推荐供应商", "推荐理由", "风险提示"]
    write_sheet(
        wb,
        "阶段性推荐",
        "阶段性推荐",
        rec_headers,
        [[row[h] for h in rec_headers] for row in recommendations],
        {"采购包": 28, "阶段性推荐供应商": 28, "推荐理由": 70, "风险提示": 80},
    )
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="分析供应商报价并生成谈判策略")
    parser.add_argument("--base-dir", default=str(PROJECT_DIR), help="项目根目录")
    parser.add_argument("--input-dir", default="采购询价本地工作流/input/supplier_quotes", help="供应商报价输入目录")
    parser.add_argument("--output-dir", default="采购询价本地工作流/output", help="输出目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir).resolve()
    input_dir = resolve_path(base_dir, args.input_dir)
    output_root = resolve_path(base_dir, args.output_dir)
    input_dir.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / f"{run_id}_供应商报价分析"
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = collect_quotes(input_dir)
    output_path = output_dir / "供应商报价分析及谈判策略.xlsx"
    write_analysis(output_path, input_dir, rows)
    log = [
        f"运行时间：{run_id}",
        f"输入目录：{input_dir}",
        f"读取报价行：{len(rows)}",
        f"输出文件：{output_path}",
    ]
    (output_dir / "处理日志.txt").write_text("\n".join(log) + "\n", encoding="utf-8")
    print("\n".join(log))


if __name__ == "__main__":
    main()
