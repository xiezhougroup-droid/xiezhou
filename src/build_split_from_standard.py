#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = TOOL_DIR.parent
DEFAULT_OUTPUT_DIR = "采购询价本地工作流/output"
DEFAULT_STANDARD_PATH = "/Users/houzhou/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/houge630973991_3046/temp/drag/永顺县芙蓉镇、青坪镇农贸市场人材机汇总表_询价清单(3).xlsx"

SOURCE_SHEET = "正式询价清单"

SOURCE_HEADERS = [
    "序号",
    "来源行号",
    "原始编号",
    "原始材料名称",
    "标准材料名称",
    "询价大类",
    "询价采购包",
    "供应商类型",
    "规格型号",
    "厚度参数",
    "参数选项",
    "单位",
    "合并数量",
    "原始参考单价",
    "原始参考合价",
    "是否需人工确认",
    "是否可直接发供应商询价",
    "询价前需补充信息",
    "备注",
    "供应商报价单价",
    "供应商报价合价",
    "品牌/厂家",
    "税率",
    "供货周期",
    "付款条件",
    "报价备注",
]

OUTPUT_HEADERS = [
    "序号",
    "来源行号",
    "追溯号",
    "原始编号",
    "原始材料名称",
    "标准材料名称",
    "询价大类",
    "询价采购包",
    "供应商类型",
    "供应方式初判",
    "规格型号",
    "厚度参数",
    "参数选项",
    "单位",
    "合并数量",
    "原始参考单价",
    "原始参考合价",
    "是否需人工确认",
    "是否可直接发供应商询价",
    "询价前需补充信息",
    "备注",
    "供应商报价单价",
    "供应商报价合价",
    "品牌/厂家",
    "税率",
    "供货周期",
    "付款条件",
    "报价备注",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def trace_id(row: dict[str, Any]) -> str:
    source_row = clean(row.get("来源行号"))
    original_code = clean(row.get("原始编号"))
    return f"{source_row}｜{original_code}" if original_code else source_row


def supply_mode(row: dict[str, Any]) -> str:
    package = clean(row.get("询价采购包"))
    supplier_type = clean(row.get("供应商类型"))
    text = f"{package} {supplier_type}"
    if "待采购员确认" in text or package in ["零星待确认包", "暂不询价"]:
        return "待确认"
    if any(keyword in text for keyword in ["门窗", "栏杆", "专业", "脚手架", "风管"]):
        return "材料供应+安装/加工边界待确认"
    if any(keyword in text for keyword in ["成套", "设备", "配电箱", "消防报警", "弱电", "灯具", "暖通设备"]):
        return "设备/材料供应"
    return "材料供应"


def read_standard_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SOURCE_SHEET]
    headers = [cell.value for cell in ws[1]]
    missing = [header for header in SOURCE_HEADERS if header not in headers]
    if missing:
        raise RuntimeError(f"样板文件字段缺失：{missing}")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if not any(clean(value) for value in values):
            continue
        row["追溯号"] = trace_id(row)
        row["供应方式初判"] = supply_mode(row)
        rows.append(row)
    return rows


def style_sheet(ws, header_row: int = 3) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    blue = PatternFill("solid", fgColor="5B9BD5")
    orange = PatternFill("solid", fgColor="FCE4D6")
    green = PatternFill("solid", fgColor="E2F0D9")
    gray = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if ws.max_column:
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = dark
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    manual_col = ready_col = package_col = None
    for idx, cell in enumerate(ws[header_row], 1):
        if cell.value == "是否需人工确认":
            manual_col = idx
        if cell.value == "是否可直接发供应商询价":
            ready_col = idx
        if cell.value == "询价采购包":
            package_col = idx

    for row in ws.iter_rows(min_row=header_row + 1):
        package_value = clean(row[package_col - 1].value) if package_col else ""
        if package_value == "暂不询价":
            fill = gray
        elif manual_col and row[manual_col - 1].value == "是":
            fill = orange
        elif ready_col and row[ready_col - 1].value == "是":
            fill = green
        else:
            fill = None
        if fill:
            for cell in row:
                cell.fill = fill

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.sheet_view.showGridLines = False
    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_titled_sheet(wb: Workbook, name: str, title: str, headers: list[str], rows: list[list[Any]], widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title)
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    for row_idx, row in enumerate(rows, 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
    style_sheet(ws)
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)


def write_output(path: Path, rows: list[dict[str, Any]], standard_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    package_counter = Counter(clean(row["询价采购包"]) for row in rows)
    category_counter = Counter(clean(row["询价大类"]) for row in rows)
    supplier_counter = Counter(clean(row["供应商类型"]) for row in rows)
    mode_counter = Counter(clean(row["供应方式初判"]) for row in rows)
    manual_rows = [row for row in rows if clean(row.get("是否需人工确认")) == "是"]
    ready_rows = [row for row in rows if clean(row.get("是否可直接发供应商询价")) == "是"]
    not_ready_rows = [row for row in rows if clean(row.get("是否可直接发供应商询价")) != "是"]

    summary_headers = ["项目", "数量/说明"]
    summary_rows = [
        ["标准来源文件", str(standard_path)],
        ["正式询价清单行数", len(rows)],
        ["询价大类数量", len(category_counter)],
        ["询价采购包数量", len(package_counter)],
        ["可直接发供应商询价", len(ready_rows)],
        ["需人工确认", len(manual_rows)],
        ["不可直接发供应商询价", len(not_ready_rows)],
        ["说明", "本文件按用户提供的昨日样板标准拆分，保留询价大类、询价采购包、供应商类型和追溯号。"],
    ]
    for package, count in package_counter.most_common():
        summary_rows.append([f"询价采购包：{package}", count])
    write_titled_sheet(wb, "处理说明", "按昨日样板标准拆分说明", summary_headers, summary_rows, {"项目": 42, "数量/说明": 96})

    widths = {
        "序号": 8,
        "来源行号": 12,
        "追溯号": 24,
        "原始编号": 20,
        "原始材料名称": 38,
        "标准材料名称": 38,
        "询价大类": 26,
        "询价采购包": 30,
        "供应商类型": 30,
        "供应方式初判": 26,
        "规格型号": 34,
        "厚度参数": 16,
        "参数选项": 42,
        "单位": 10,
        "合并数量": 14,
        "原始参考单价": 14,
        "原始参考合价": 16,
        "是否需人工确认": 16,
        "是否可直接发供应商询价": 20,
        "询价前需补充信息": 52,
        "备注": 42,
        "供应商报价单价": 16,
        "供应商报价合价": 16,
        "品牌/厂家": 16,
        "税率": 12,
        "供货周期": 14,
        "付款条件": 18,
        "报价备注": 20,
    }
    matrix = [[row.get(header, "") for header in OUTPUT_HEADERS] for row in rows]
    write_titled_sheet(wb, "采购包拆分总表", "采购包拆分总表（按昨日样板标准）", OUTPUT_HEADERS, matrix, widths)
    write_titled_sheet(wb, "可直接询价项", "可直接询价项", OUTPUT_HEADERS, [[row.get(header, "") for header in OUTPUT_HEADERS] for row in ready_rows], widths)
    write_titled_sheet(wb, "需人工确认项", "需人工确认项", OUTPUT_HEADERS, [[row.get(header, "") for header in OUTPUT_HEADERS] for row in manual_rows], widths)
    write_titled_sheet(wb, "不可直接询价项", "不可直接询价项", OUTPUT_HEADERS, [[row.get(header, "") for header in OUTPUT_HEADERS] for row in not_ready_rows], widths)

    package_headers = ["询价采购包", "询价大类", "供应商类型", "供应方式初判", "行数", "可直接询价", "需人工确认", "不可直接询价"]
    package_rows = []
    for package, count in package_counter.most_common():
        items = [row for row in rows if clean(row["询价采购包"]) == package]
        package_rows.append([
            package,
            Counter(clean(row["询价大类"]) for row in items).most_common(1)[0][0],
            Counter(clean(row["供应商类型"]) for row in items).most_common(1)[0][0],
            Counter(clean(row["供应方式初判"]) for row in items).most_common(1)[0][0],
            count,
            sum(1 for row in items if clean(row.get("是否可直接发供应商询价")) == "是"),
            sum(1 for row in items if clean(row.get("是否需人工确认")) == "是"),
            sum(1 for row in items if clean(row.get("是否可直接发供应商询价")) != "是"),
        ])
    write_titled_sheet(
        wb,
        "采购包汇总",
        "采购包汇总",
        package_headers,
        package_rows,
        {"询价采购包": 34, "询价大类": 30, "供应商类型": 34, "供应方式初判": 28, "行数": 10, "可直接询价": 14, "需人工确认": 14, "不可直接询价": 16},
    )

    category_headers = ["询价大类", "行数", "可直接询价", "需人工确认", "主要采购包"]
    category_rows = []
    for category, count in category_counter.most_common():
        items = [row for row in rows if clean(row["询价大类"]) == category]
        category_rows.append([
            category,
            count,
            sum(1 for row in items if clean(row.get("是否可直接发供应商询价")) == "是"),
            sum(1 for row in items if clean(row.get("是否需人工确认")) == "是"),
            "、".join(package for package, _ in Counter(clean(row["询价采购包"]) for row in items).most_common(5)),
        ])
    write_titled_sheet(
        wb,
        "询价大类汇总",
        "询价大类汇总",
        category_headers,
        category_rows,
        {"询价大类": 32, "行数": 10, "可直接询价": 14, "需人工确认": 14, "主要采购包": 90},
    )

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按昨日样板标准生成采购包拆分总表")
    parser.add_argument("--standard", default=DEFAULT_STANDARD_PATH, help="昨日样板标准 xlsx 路径")
    parser.add_argument("--base-dir", default=str(PROJECT_DIR), help="项目根目录")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="输出目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir).resolve()
    standard_path = resolve_path(base_dir, args.standard)
    output_root = resolve_path(base_dir, args.output_dir)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / f"{run_id}_按昨日样板标准拆分"
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = read_standard_rows(standard_path)
    output_path = output_dir / "采购包拆分总表-按昨日样板标准.xlsx"
    write_output(output_path, rows, standard_path)

    log_lines = [
        f"运行时间：{run_id}",
        f"标准来源文件：{standard_path}",
        f"输出目录：{output_dir}",
        f"正式询价清单行数：{len(rows)}",
        f"询价采购包数量：{len(set(clean(row['询价采购包']) for row in rows))}",
        f"可直接询价项：{sum(1 for row in rows if clean(row.get('是否可直接发供应商询价')) == '是')}",
        f"需人工确认项：{sum(1 for row in rows if clean(row.get('是否需人工确认')) == '是')}",
        f"输出文件：{output_path}",
    ]
    (output_dir / "处理日志.txt").write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    print("\n".join(log_lines))


if __name__ == "__main__":
    main()
