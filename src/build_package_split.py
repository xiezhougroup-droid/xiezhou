#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = TOOL_DIR.parent
CONFIG_DIR = TOOL_DIR / "config"

SOURCE_HEADERS = [
    "原始行号",
    "原始编号",
    "原始材料名称",
    "标准材料名称",
    "规格型号",
    "单位",
    "数量",
    "参考单价",
    "参考合价",
    "原始类别",
    "询价状态",
    "材料大类",
    "建议供应商包",
    "需人工确认",
    "确认原因",
]

OUTPUT_HEADERS = [
    "原始行号",
    "追溯号",
    "原始编号",
    "原始材料名称",
    "标准材料名称",
    "规格型号",
    "单位",
    "数量",
    "材料大类",
    "原建议供应商包",
    "智能识别采购包",
    "供应方式",
    "建议询价模板类型",
    "是否适合直接询价",
    "需人工确认",
    "确认原因",
    "规则命中",
    "识别置信度",
]


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def resolve_path(base_dir: Path, configured: str) -> Path:
    path = Path(configured)
    return path if path.is_absolute() else base_dir / path


def trace_id(row: dict[str, Any]) -> str:
    original_row = clean(row.get("原始行号"))
    original_code = clean(row.get("原始编号"))
    return f"{original_row}｜{original_code}" if original_code else original_row


def row_to_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}


def text_for_match(row: dict[str, Any]) -> str:
    return " ".join(
        clean(row.get(field))
        for field in ["原始材料名称", "标准材料名称", "规格型号", "材料大类", "建议供应商包"]
    )


def any_keyword(text: str, keywords: list[str]) -> bool:
    return any(keyword and keyword in text for keyword in keywords)


def rule_matches(row: dict[str, Any], rule: dict[str, Any]) -> bool:
    source_package = clean(row.get("建议供应商包"))
    text = text_for_match(row)
    source_packages = rule.get("source_supplier_packages", [])
    include_keywords = rule.get("include_keywords", [])
    exclude_keywords = rule.get("exclude_keywords", [])

    if source_packages and source_package not in source_packages:
        return False
    if include_keywords and not any_keyword(text, include_keywords):
        return False
    if exclude_keywords and any_keyword(text, exclude_keywords):
        return False
    return True


def fallback_classification(row: dict[str, Any], cfg: dict[str, Any]) -> dict[str, str]:
    source_package = clean(row.get("建议供应商包"))
    fallback = cfg.get("fallback_by_source_package", {}).get(source_package)
    if fallback:
        return {
            "智能识别采购包": fallback[0],
            "建议询价模板类型": fallback[1],
            "供应方式": fallback[2],
            "规则命中": f"fallback_by_source_package:{source_package}",
            "识别置信度": "中" if source_package != "其他材料供应商包" else "低",
            "默认就绪状态": "可直接询价" if source_package in cfg.get("direct_quote_source_packages", []) else "需补充后询价",
            "默认确认原因": "" if source_package in cfg.get("direct_quote_source_packages", []) else "当前仅按原建议供应商包归类，需采购员确认是否适合直接询价。",
        }
    return {
        "智能识别采购包": "未识别采购包",
        "建议询价模板类型": "待确认",
        "供应方式": "待确认",
        "规则命中": "unmatched",
        "识别置信度": "低",
        "默认就绪状态": "需补充后询价",
        "默认确认原因": "未匹配到明确采购包规则，需人工分类。",
    }


def classify(row: dict[str, Any], cfg: dict[str, Any]) -> dict[str, str]:
    for rule in cfg.get("rules", []):
        if rule_matches(row, rule):
            return {
                "智能识别采购包": rule["package_name"],
                "建议询价模板类型": rule["template_type"],
                "供应方式": rule["supply_mode"],
                "规则命中": f"rule:{rule['key']}",
                "识别置信度": rule.get("confidence", "中"),
                "默认就绪状态": rule.get("quote_readiness", "可直接询价"),
                "默认确认原因": rule.get("manual_reason", ""),
            }
    return fallback_classification(row, cfg)


def needs_manual(row: dict[str, Any], classification: dict[str, str], cfg: dict[str, Any]) -> tuple[str, str, str]:
    reasons: list[str] = []
    original_manual = clean(row.get("需人工确认"))
    original_reason = clean(row.get("确认原因"))
    spec = clean(row.get("规格型号"))
    name = clean(row.get("标准材料名称")) or clean(row.get("原始材料名称"))
    text = text_for_match(row)

    if original_manual in cfg.get("manual_confirm_values", []):
        reasons.append(original_reason or "底表已标记需人工确认")

    if not spec and any_keyword(text, cfg.get("missing_spec_sensitive_keywords", [])):
        reasons.append("规格敏感材料缺少明确规格型号")

    if spec and any(value in spec for value in cfg.get("weak_spec_values", [])):
        reasons.append("规格描述为综合/各规格/配套，需确认报价口径")

    if classification["识别置信度"] == "低":
        reasons.append(classification["默认确认原因"])

    if "待确认" in classification["供应方式"]:
        reasons.append(classification["默认确认原因"] or "供应方式或范围边界待确认")

    if classification["默认就绪状态"] != "可直接询价":
        reasons.append(classification["默认确认原因"] or classification["默认就绪状态"])

    if not clean(row.get("单位")) or not clean(row.get("数量")):
        reasons.append("缺少单位或数量，暂不适合直接询价")

    reasons = list(dict.fromkeys(reason for reason in reasons if reason))
    manual = "是" if reasons else "否"
    ready = "否" if reasons else "是"
    return manual, ready, "；".join(reasons)


def read_boq(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if headers[: len(SOURCE_HEADERS)] != SOURCE_HEADERS:
        raise RuntimeError("底表字段与预期不一致，请检查正式询价清单底表。")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = row_to_dict(SOURCE_HEADERS, values)
        if clean(row.get("询价状态")) == "拟询价":
            rows.append(row)
    return rows


def build_output_rows(rows: list[dict[str, Any]], cfg: dict[str, Any]) -> list[dict[str, Any]]:
    output_rows: list[dict[str, Any]] = []
    for row in rows:
        classification = classify(row, cfg)
        manual, ready, reason = needs_manual(row, classification, cfg)
        output_rows.append({
            "原始行号": clean(row.get("原始行号")),
            "追溯号": trace_id(row),
            "原始编号": clean(row.get("原始编号")),
            "原始材料名称": clean(row.get("原始材料名称")),
            "标准材料名称": clean(row.get("标准材料名称")),
            "规格型号": clean(row.get("规格型号")),
            "单位": clean(row.get("单位")),
            "数量": row.get("数量"),
            "材料大类": clean(row.get("材料大类")),
            "原建议供应商包": clean(row.get("建议供应商包")),
            "智能识别采购包": classification["智能识别采购包"],
            "供应方式": classification["供应方式"],
            "建议询价模板类型": classification["建议询价模板类型"],
            "是否适合直接询价": ready,
            "需人工确认": manual,
            "确认原因": reason,
            "规则命中": classification["规则命中"],
            "识别置信度": classification["识别置信度"],
        })
    return output_rows


def style_sheet(ws, freeze_row: int = 3) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    blue = PatternFill("solid", fgColor="5B9BD5")
    orange = PatternFill("solid", fgColor="FCE4D6")
    green = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if ws.max_column:
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = dark
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    manual_col = None
    ready_col = None
    for idx, cell in enumerate(ws[3], 1):
        if cell.value == "需人工确认":
            manual_col = idx
        if cell.value == "是否适合直接询价":
            ready_col = idx
    for row in ws.iter_rows(min_row=4):
        if manual_col and row[manual_col - 1].value == "是":
            for cell in row:
                cell.fill = orange
        elif ready_col and row[ready_col - 1].value == "是":
            for cell in row:
                cell.fill = green

    ws.freeze_panes = ws.cell(freeze_row + 1, 1)
    ws.sheet_view.showGridLines = False
    if ws.max_row >= freeze_row and ws.max_column:
        ws.auto_filter.ref = f"A{freeze_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def set_widths(ws, widths: dict[str, int]) -> None:
    for idx, cell in enumerate(ws[3], 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(clean(cell.value), 16)


def write_titled_sheet(wb: Workbook, name: str, title: str, headers: list[str], rows: list[list[Any]], widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title)
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    style_sheet(ws)
    set_widths(ws, widths)


def write_workbook(path: Path, rows: list[dict[str, Any]], source_rows_count: int) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    by_package = Counter(row["智能识别采购包"] for row in rows)
    by_mode = Counter(row["供应方式"] for row in rows)
    by_template = Counter(row["建议询价模板类型"] for row in rows)
    manual_rows = [row for row in rows if row["需人工确认"] == "是"]
    ready_rows = [row for row in rows if row["是否适合直接询价"] == "是"]

    summary_headers = ["项目", "数量/说明"]
    summary_rows = [
        ["底表拟询价项", source_rows_count],
        ["拆分后采购包数量", len(by_package)],
        ["适合直接询价项", len(ready_rows)],
        ["需人工确认项", len(manual_rows)],
        ["说明", "本表为采购包智能拆分原型，不修改原始底表。后续请优先校准规则，再批量生成各包邀请函。"],
    ]
    for package, count in by_package.most_common():
        summary_rows.append([f"采购包：{package}", count])
    for mode, count in by_mode.most_common():
        summary_rows.append([f"供应方式：{mode}", count])
    for template, count in by_template.most_common():
        summary_rows.append([f"模板类型：{template}", count])
    write_titled_sheet(wb, "处理说明", "采购包拆分处理说明", summary_headers, summary_rows, {"项目": 42, "数量/说明": 72})

    widths = {
        "原始行号": 12,
        "追溯号": 22,
        "原始编号": 20,
        "原始材料名称": 34,
        "标准材料名称": 34,
        "规格型号": 30,
        "单位": 10,
        "数量": 12,
        "材料大类": 18,
        "原建议供应商包": 24,
        "智能识别采购包": 30,
        "供应方式": 24,
        "建议询价模板类型": 26,
        "是否适合直接询价": 16,
        "需人工确认": 14,
        "确认原因": 56,
        "规则命中": 30,
        "识别置信度": 14,
    }
    matrix = [[row[h] for h in OUTPUT_HEADERS] for row in rows]
    write_titled_sheet(wb, "采购包拆分总表", "采购包拆分总表", OUTPUT_HEADERS, matrix, widths)
    write_titled_sheet(wb, "可直接询价项", "可直接询价项", OUTPUT_HEADERS, [[row[h] for h in OUTPUT_HEADERS] for row in ready_rows], widths)
    write_titled_sheet(wb, "需人工确认项", "需人工确认项", OUTPUT_HEADERS, [[row[h] for h in OUTPUT_HEADERS] for row in manual_rows], widths)

    package_headers = ["智能识别采购包", "项数", "可直接询价", "需人工确认", "供应方式", "建议询价模板类型"]
    package_rows = []
    for package, count in by_package.most_common():
        package_items = [row for row in rows if row["智能识别采购包"] == package]
        package_rows.append([
            package,
            count,
            sum(1 for row in package_items if row["是否适合直接询价"] == "是"),
            sum(1 for row in package_items if row["需人工确认"] == "是"),
            Counter(row["供应方式"] for row in package_items).most_common(1)[0][0],
            Counter(row["建议询价模板类型"] for row in package_items).most_common(1)[0][0],
        ])
    write_titled_sheet(
        wb,
        "采购包汇总",
        "采购包汇总",
        package_headers,
        package_rows,
        {"智能识别采购包": 36, "项数": 10, "可直接询价": 14, "需人工确认": 14, "供应方式": 28, "建议询价模板类型": 30},
    )

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成采购包拆分总表")
    parser.add_argument("--base-dir", default=str(PROJECT_DIR), help="项目根目录")
    parser.add_argument("--config", default=str(CONFIG_DIR / "procurement_classification_rules.json"), help="采购包拆分规则配置")
    parser.add_argument("--boq", default=None, help="临时指定底表路径")
    parser.add_argument("--output-dir", default=None, help="临时指定输出目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir).resolve()
    cfg = load_json(Path(args.config))
    boq_path = resolve_path(base_dir, args.boq or cfg["boq_file"])
    output_root = resolve_path(base_dir, args.output_dir or cfg["default_output_dir"])
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / f"{run_id}_采购包拆分总表"
    output_dir.mkdir(parents=True, exist_ok=True)

    source_rows = read_boq(boq_path, cfg["boq_sheet"])
    output_rows = build_output_rows(source_rows, cfg)
    output_path = output_dir / "采购包智能拆分总表.xlsx"
    write_workbook(output_path, output_rows, len(source_rows))

    summary = [
        f"运行时间：{run_id}",
        f"底表：{boq_path}",
        f"输出目录：{output_dir}",
        f"拟询价项：{len(source_rows)}",
        f"采购包数量：{len(set(row['智能识别采购包'] for row in output_rows))}",
        f"可直接询价项：{sum(1 for row in output_rows if row['是否适合直接询价'] == '是')}",
        f"需人工确认项：{sum(1 for row in output_rows if row['需人工确认'] == '是')}",
        f"输出文件：{output_path}",
    ]
    (output_dir / "处理日志.txt").write_text("\n".join(summary) + "\n", encoding="utf-8")
    print("\n".join(summary))


if __name__ == "__main__":
    main()
