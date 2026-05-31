#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

CODEX_PYTHON_PACKAGES = Path("/Users/houzhou/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages")
if CODEX_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_PYTHON_PACKAGES))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = TOOL_DIR.parent

FEEDBACK_HEADERS = [
    "采购包",
    "供应商",
    "最终报价总价",
    "谈判后付款条件",
    "谈判后供货周期",
    "技术/品牌确认情况",
    "采购员确认状态",
    "推荐意见",
    "风险备注",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def style_sheet(ws, header_row=3) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    blue = PatternFill("solid", fgColor="5B9BD5")
    green = PatternFill("solid", fgColor="E2F0D9")
    orange = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if ws.max_row >= header_row:
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1):
        text = " ".join(clean(cell.value) for cell in row)
        if "推荐" in text and "不推荐" not in text:
            for cell in row:
                cell.fill = green
        elif "待确认" in text or "风险" in text or "不推荐" in text:
            for cell in row:
                cell.fill = orange
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.sheet_view.showGridLines = False
    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_sheet(wb: Workbook, name: str, title: str, headers: list[str], rows: list[list[Any]], widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws.cell(1, 1, title)
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    style_sheet(ws)
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)


def create_feedback_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "谈判反馈表",
        "谈判反馈表",
        FEEDBACK_HEADERS,
        [
            ["示例采购包", "示例供应商", "", "待填写", "待填写", "待填写", "待确认/已确认", "推荐/备选/不推荐", "待填写"],
        ],
        {"采购包": 28, "供应商": 28, "最终报价总价": 16, "谈判后付款条件": 32, "谈判后供货周期": 22, "技术/品牌确认情况": 36, "采购员确认状态": 20, "推荐意见": 18, "风险备注": 60},
    )
    wb.save(path)


def read_feedback(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "谈判反馈表" not in wb.sheetnames:
        return []
    ws = wb["谈判反馈表"]
    headers = [cell.value for cell in ws[3]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=4, values_only=True):
        row = dict(zip(headers, values))
        if clean(row.get("采购包")) and clean(row.get("供应商")) and clean(row.get("采购包")) != "示例采购包":
            rows.append(row)
    return rows


def build_recommendations(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    packages = sorted({clean(row.get("采购包")) for row in rows})
    for package in packages:
        items = [row for row in rows if clean(row.get("采购包")) == package]
        confirmed = [row for row in items if "已确认" in clean(row.get("采购员确认状态"))]
        recommended = [row for row in confirmed if clean(row.get("推荐意见")) == "推荐"]
        candidates = recommended or sorted(confirmed, key=lambda r: number(r.get("最终报价总价")) or 10**18)
        if candidates:
            chosen = candidates[0]
            status = "推荐合作"
            reason = f"采购员已确认；最终报价 {clean(chosen.get('最终报价总价')) or '待补充'}；付款条件：{clean(chosen.get('谈判后付款条件')) or '待补充'}；供货周期：{clean(chosen.get('谈判后供货周期')) or '待补充'}。"
            risk = clean(chosen.get("风险备注")) or "建议合同中锁定品牌、规格、税率、供货周期、付款条件及偏离事项。"
            supplier = clean(chosen.get("供应商"))
        else:
            status = "暂不推荐"
            reason = "尚未收到采购员对该采购包的完整确认，或供应商谈判反馈未完成。"
            risk = "需补齐最终报价、商务条件、技术/品牌确认情况及采购员确认状态。"
            supplier = ""
        result.append({
            "采购包": package,
            "推荐状态": status,
            "推荐供应商": supplier,
            "推荐理由": reason,
            "风险及合同注意事项": risk,
        })
    return result


def write_final(output_path: Path, feedback_path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    if not rows:
        write_sheet(
            wb,
            "使用说明",
            "最终推荐使用说明",
            ["事项", "说明"],
            [
                ["当前状态", "未读取到有效谈判反馈。"],
                ["反馈表路径", str(feedback_path)],
                ["填写要求", "在“谈判反馈表”中填写采购包、供应商、最终报价总价、付款条件、供货周期、技术/品牌确认情况、采购员确认状态和推荐意见。"],
                ["确认口径", "只有采购员确认状态包含“已确认”的记录，才会进入最终推荐判断。"],
            ],
            {"事项": 28, "说明": 100},
        )
        wb.save(output_path)
        return
    recs = build_recommendations(rows)
    write_sheet(
        wb,
        "最终推荐结果",
        "最终推荐结果",
        ["采购包", "推荐状态", "推荐供应商", "推荐理由", "风险及合同注意事项"],
        [[row["采购包"], row["推荐状态"], row["推荐供应商"], row["推荐理由"], row["风险及合同注意事项"]] for row in recs],
        {"采购包": 28, "推荐状态": 16, "推荐供应商": 28, "推荐理由": 80, "风险及合同注意事项": 90},
    )
    write_sheet(
        wb,
        "谈判反馈明细",
        "谈判反馈明细",
        FEEDBACK_HEADERS,
        [[row.get(h, "") for h in FEEDBACK_HEADERS] for row in rows],
        {"采购包": 28, "供应商": 28, "最终报价总价": 16, "谈判后付款条件": 32, "谈判后供货周期": 22, "技术/品牌确认情况": 36, "采购员确认状态": 20, "推荐意见": 18, "风险备注": 60},
    )
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据采购员谈判反馈生成最终推荐结果")
    parser.add_argument("--base-dir", default=str(PROJECT_DIR), help="项目根目录")
    parser.add_argument("--feedback", default="采购询价本地工作流/input/negotiation_feedback.xlsx", help="谈判反馈表")
    parser.add_argument("--output-dir", default="采购询价本地工作流/output", help="输出目录")
    parser.add_argument("--package", default="", help="只生成指定采购包的最终推荐结果；为空则生成全部采购包")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir).resolve()
    feedback_path = resolve_path(base_dir, args.feedback)
    output_root = resolve_path(base_dir, args.output_dir)
    feedback_path.parent.mkdir(parents=True, exist_ok=True)
    if not feedback_path.exists():
        create_feedback_template(feedback_path)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_filter = clean(args.package)
    output_dir_name = f"{run_id}_{package_filter}_最终推荐结果" if package_filter else f"{run_id}_最终推荐结果"
    output_dir = output_root / output_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = read_feedback(feedback_path)
    if package_filter:
        rows = [row for row in rows if clean(row.get("采购包")) == package_filter]
    output_path = output_dir / "采购推荐合作结果.xlsx"
    write_final(output_path, feedback_path, rows)
    log = [
        f"运行时间：{run_id}",
        f"谈判反馈表：{feedback_path}",
        f"采购包筛选：{package_filter or '全部采购包'}",
        f"有效反馈记录：{len(rows)}",
        f"输出文件：{output_path}",
    ]
    (output_dir / "处理日志.txt").write_text("\n".join(log) + "\n", encoding="utf-8")
    print("\n".join(log))


if __name__ == "__main__":
    main()
